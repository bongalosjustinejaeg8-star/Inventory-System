import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

Inventory_Database = "product.xlsx"
Sales_Database = "sale.xlsx"
user_database = "user.xlsx"
Movement_Database = "inventory_movements.xlsx"

def create_if_not_exists(filename, headers):
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(filename)

create_if_not_exists(Inventory_Database, ["product_id","product_name","category","price","stock_quantity","reorder_level"])
create_if_not_exists(Sales_Database, ["sale_id","date","product_id","quantity","unit_price","total"])
create_if_not_exists(user_database, ["username","password","role"])
create_if_not_exists(Movement_Database, ["movement_id","product_id","movement_type","quantity","date","remarks"])

wb1 = load_workbook(Inventory_Database)
ws1 = wb1.active
wb2 = load_workbook(Sales_Database)
ws2 = wb2.active
wb3 = load_workbook(user_database)
ws3 = wb3.active
wb4 = load_workbook(Movement_Database)
ws4 = wb4.active

def safe_int(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0

def safe_float(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0

def log_movement(product_id, movement_type, quantity, remarks):
    movement_id = ws4.max_row
    date_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws4.append([movement_id, product_id, movement_type, quantity, date_now, remarks])
    wb4.save(Movement_Database)

def add_new(product_Id, name, category, price, qty, rodlvl):
    ws1.append([product_Id.upper(), name.title(), category.title(), safe_float(price), safe_int(qty), safe_int(rodlvl)])
    wb1.save(Inventory_Database)
    log_movement(product_Id.upper(), "IN", qty, "Initial stock")

def change_stock(product_id, new_stock):
    for row in ws1.iter_rows(min_row=2):
        if str(row[0].value).strip().upper() == product_id.upper():
            old_stock = safe_int(row[4].value)
            diff = safe_int(new_stock) - old_stock
            row[4].value = safe_int(new_stock)
            wb1.save(Inventory_Database)
            movement_type = "IN" if diff > 0 else "OUT"
            log_movement(product_id, movement_type, abs(diff), "Manual stock change")
            return True
    return False

def get_price(product_id):
    for row in ws1.iter_rows(min_row=2, values_only=True):
        if str(row[0]).strip().upper() == product_id.upper():
            return safe_float(row[3])
    return None

def update_stock(product_id, qty_sold):
    for row in ws1.iter_rows(min_row=2):
        if str(row[0].value).strip().upper() == product_id.upper():
            current_stock = safe_int(row[4].value)
            new_stock = current_stock - safe_int(qty_sold)
            if new_stock < 0:
                print(f"Not enough stock for {product_id}! Only {current_stock} left.")
                return False
            row[4].value = new_stock
            wb1.save(Inventory_Database)
            log_movement(product_id, "OUT", qty_sold, "Stock adjustment")
            return True
    return False

def buy(product_id, qty, sale_id):
    if not update_stock(product_id, qty):
        print("Sale failed. Stock not updated.")
        return False
    price = get_price(product_id)
    subtotal = safe_float(price) * safe_int(qty)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws2.append([sale_id, now, product_id, qty, price, subtotal])
    wb2.save(Sales_Database)
    log_movement(product_id, "SALE", qty, "Customer purchase")
    return True

def product_name(product_id):
    for row in ws1.iter_rows(min_row=2, values_only=True):
        if str(row[0]).strip().upper() == product_id.upper():
            return row[1] or "-"
    return "-"

def print_receipt(sales_id):
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[0] == safe_int(sales_id):
            return row[2], row[3], row[4], row[5]
    return None

def save():
    wb1.save("product.xlsx")
    wb2.save("sale.xlsx")
    wb3.save("user.xlsx")
    wb4.save("inventory_movements.xlsx")

def list_products():
    print("\n====== INVENTORY LIST ======")
    print(f"{'ID':<10} {'Name':<20} {'Category':<15} {'Price':<10} {'Stock':<10} {'Reorder':<10}")
    
    for row in ws1.iter_rows(min_row=2, values_only=True):
        pid, name, cat, price, stock, reorder = row
        print(f"{pid or '-':<10} {name or '-':<20} {cat or '-':<15} {safe_float(price):<10} {safe_int(stock):<10} {safe_int(reorder):<10}")

def list_sales():
    print("\n====== SALES RECORDS ======")
    print(f"{'Sale ID':<10} {'Date':<20} {'Product ID':<10} {'Qty':<5} {'Unit Price':<10} {'Total':<10}")
    print("-"*70)
    for row in ws2.iter_rows(min_row=2, values_only=True):
        sid, date, pid, qty, price, total = row
        print(f"{sid or '-':<10} {date or '-':<20} {pid or '-':<10} {safe_int(qty):<5} {safe_float(price):<10} {safe_float(total):<10}")

def list_inventory_movements():
    print("\n====== INVENTORY MOVEMENTS ======")
    print(f"{'ID':<5} {'Product ID':<10} {'Type':<10} {'Qty':<5} {'Date':<20} {'Remarks':<20}")
    print("-"*75)
    for row in ws4.iter_rows(min_row=2, values_only=True):
        mid, pid, mtype, qty, date, remarks = row
        print(f"{mid or '-':<5} {pid or '-':<10} {mtype or '-':<10} {safe_int(qty):<5} {date or '-':<20} {remarks or '-':<20}")

def sales_summary():
    total_sales = total_items = total_revenue = 0
    for row in ws2.iter_rows(min_row=2, values_only=True):
        total_sales += 1 if row[0] is not None else 0
        total_items += safe_int(row[3])
        total_revenue += safe_float(row[5])
    print("\n====== SALES SUMMARY ======")
    print(f"Total Sales Transactions: {total_sales}")
    print(f"Total Items Sold       : {total_items}")
    print(f"Total Revenue          : {total_revenue}")

def best_selling_products():
    product_sales = {}
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            pid = row[2]
            qty = row[3] if row[3] is not None else 0
            if pid in product_sales:
                product_sales[pid] += qty
            else:
                product_sales[pid] = qty

    sorted_sales = sorted(product_sales.items(), key=lambda x: x[1], reverse=True)

    print("\n====== BEST-SELLING PRODUCTS ======")
    print(f"{'Product ID':<10} {'Product Name':<20} {'Total Sold':<10}")
    print("-"*45)

    for pid, total in sorted_sales[:10]: 
        name = product_name(pid)
        if name is None:
            name = "-"  
        pid = str(pid)
        total = total if total is not None else 0
        print(f"{pid:<10} {name:<20} {total:<10}")


def low_stock_alert():
    print("\n====== LOW STOCK ALERTS ======")
    print(f"{'Product ID':<10} {'Name':<20} {'Stock':<10} {'Reorder':<10}")
    low_stock_found = False
    for row in ws1.iter_rows(min_row=2, values_only=True):
        pid, name, cat, price, stock, reorder = row

    
        pid = str(pid) if pid is not None else "-"
        name = name if name is not None else "-"
        stock = safe_int(stock)
        reorder = safe_int(reorder)

        if stock <= reorder:
            low_stock_found = True
            print(f"{pid:<10} {name:<20} {stock:<10} {reorder:<10}")

    if not low_stock_found:
        print("All products have sufficient stock.")

# ----------------- TERMINAL MENU -----------------
def main_menu():
    while True:
        print("\n====== SALES & INVENTORY SYSTEM ======")
        print("1. Add New Product")
        print("2. Buy Product")
        print("3. Change Stock (Manual)")
        print("4. Check Product Price")
        print("5. Print Receipt")
        print("6. List All Products")
        print("7. List All Sales")
        print("8. List Inventory Movements")
        print("9. Sales Summary")
        print("10. Best-Selling Products")
        print("11. Low-Stock Alerts")
        print("12. Exit")


        choice = input("Select option: ")

        if choice == "1":
            pid = input("Product ID: ").upper()
            name = input("Product Name: ")
            cat = input("Category: ")
            price = float(input("Price: "))
            qty = int(input("Quantity: "))
            reorder = int(input("Reorder Level: "))
            add_new(pid, name, cat, price, qty, reorder)
            print("Product added successfully")

        elif choice == "2":
            pid = input("Product ID: ").upper()
            qty = int(input("Quantity to buy: "))
            sale_id = ws2.max_row
            if buy(pid, qty, sale_id):
                print("Purchase successful")

        elif choice == "3":
            pid = input("Product ID: ").upper()
            new_stock = int(input("New Stock Quantity: "))
            if change_stock(pid, new_stock):
                print("Stock updated")

        elif choice == "4":
            pid = input("Product ID: ").upper()
            price = get_price(pid)
            if price is not None:
                print(f"Price: {price}")
            else:
                print("Product not found")

        elif choice == "5":
            sale_id = input("Enter Sale ID: ")
            receipt = print_receipt(sale_id)
            if receipt:
                print("RECEIPT")
                print(f"Product ID : {receipt[0]}")
                print(f"Quantity   : {receipt[1]}")
                print(f"Unit Price : {receipt[2]}")
                print(f"Total      : {receipt[3]}")
            else:
                print("Sale not found")

        elif choice == "6":
            list_products()
        elif choice == "7":
            list_sales()
        elif choice == "8":
            list_inventory_movements()
        elif choice == "9":
            sales_summary()
        elif choice == "10":
            best_selling_products()
        elif choice == "11":
            low_stock_alert()
        elif choice == "12":
            save()
            print("Exiting system...")
            break
        else:
            print("Invalid option")

# ----------------- RUN SYSTEM -----------------
main_menu()
