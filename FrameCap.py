from openpyxl import load_workbook, Workbook

filename = "Database.xlsx"

# Try to load the workbook, create if it doesn't exist
try:
    wb = load_workbook(filename)
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    ws.append(["Product ID", "Product Name", "Category", "Price", "Stock Quantity", "Reorder Level" ])  # headers


# Read and display stored data
def print_inv():
    print("\nCurrent Inventory:")
    for product, qty in ws.iter_rows(min_row=2, values_only=True):
        print(product, qty)


def add_new():
    Id = input("Enter Product ID: ").upper()
    name = input("Enter Product Name: ").title()
    category = input("Enter Product Category: ").title()
    price = int(input("Enter Price: "))
    qty = int(input("Enter Stocks: "))
    rodlvl = int(input("Enter Reorder Level: "))
    ws.append([Id, name,category,price,qty,rodlvl])
    wb.save("Database.xlsx")
    print(f"Added {name} with {qty} stocks")


def change_stock():
    product_id = input("enter product ID: ").upper()
    found = False

    for rows in ws.iter_rows(min_row=2):
        if rows[0].value == product_id:
            found = True
    
    if not found: 
        print("item not found")
        descicion = input("press any character to exit or A to add new product: ").upper()
        if descicion == "A":
            add_new()

    quantity = int(input("enter new quantity: "))
      
    for rows in ws.iter_rows(min_row=2):
        if rows[0].value == product_id:
            rows[5].value = quantity
            print(f"{product_id} stocks updated to {quantity}")
            wb.save("DatabaseCap.xlsx")
            found = True
            break

def remv_item():
    item = input("Enter Product ID to remove: ").upper()
    found = False
    for row in ws.iter_rows(min_row=2):
        if row[0].value == item:
            decision = input(f"are you sure you want to remove {row[1].value}? (y/n): ").lower()
            if decision == "y":
                print(f"{row[0].value} removed")
                row[0].value = None
                row[1].value = None
                break
            else:
                print("removal cancelled")
                break
    if not found:
        d = input("item id not found, wanna try again? y/n: ").lower()
        if d == "y":
            remv_item()






while True:
    print("\nInventory Management System")
    print("1. Add New Product")
    print("2. Change Stock Quantity")
    print("3. Remove Product")
    print("4. View Inventory")
    print("5. Exit")
    choice = input("Select an option (1-5): ")

    if choice == '1':
        add_new()        
    elif choice == '2':
        change_stock()
    elif choice == '3':
        remv_item()
    elif choice == '4':
        print_inv()
    elif choice == '5':
        print("Exiting the program.")
        break
    else:
        print("Invalid choice. Please select a valid option.")

# Save the workbook
wb.save(filename)

