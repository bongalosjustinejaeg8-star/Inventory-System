import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

# ADDED MOVMENT DATA BASE
Inventory_Database = "Database.xlsx"
Sales_Database = "sale.xlsx"
User_Database = "user.xlsx"
Movement_Database = "inventory_movements.xlsx"

# -ADDED CREATE FILES IF NOT EXIST
def create_if_not_exists(filename, headers):
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(filename)

create_if_not_exists(Inventory_Database, ["product_id","product_name","category","price","stock_quantity","reorder_level"])
create_if_not_exists(Sales_Database, ["sale_id","date","product_id","quantity","unit_price","total"])
create_if_not_exists(User_Database, ["username","password","role"])
create_if_not_exists(Movement_Database, ["movement_id","product_id","movement_type","quantity","date","remarks"])


wb1 = load_workbook(Inventory_Database)
ws1 = wb1.active
wb2 = load_workbook(Sales_Database)
ws2 = wb2.active
wb3 = load_workbook(User_Database)
ws3 = wb3.active
wb4 = load_workbook(Movement_Database)
ws4 = wb4.active

# ADDED ERROR HANDLING
def safe_int(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0

def safe_float(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0

# FIXED SPELLING HAHAHAHAHHA AND ADDED MOVEMENT LOG
def change_stock(product_id, new_stock):
    for row in ws1.iter_rows(min_row=2):
        if str(row[0].value).strip().upper() == product_id.upper():
            old_stock = safe_int(row[4].value)
            diff = safe_int(new_stock) - old_stock
            row[4].value = safe_int(new_stock)
            wb1.save(Inventory_Database)
            movement_type = "IN" if diff > 0 else "OUT"
            log_movement(product_id, movement_type, abs(diff), "Manual stock change")
            return True
    return False

# ADDED .UPPER FOR SAFETY AND CHANGE "NONE" TO "-" FOR NEATNESS
def product_name(product_id):
    for row in ws1.iter_rows(min_row=2, values_only=True):
        if str(row[0]).strip().upper() == product_id.upper():
            return row[1] or "-"
    return "-"


def login(user,password):
    for row in ws3.iter_rows(min_row = 1,values_only = True):
        if row[0] == user and row[2] == password:
            return True
    return False
    
def checkadmin(user):
    for row in ws3.iter_rows(min_row = 1,values_only = True):
        if row[0] == user:
            if row[1] == "Admin":
                return True
            return False
    return False
def checkaudit(user):
    for row in ws3.iter_rows(min_row = 1,values_only = True):
        if row[0] == user:
            if row[1] == "Audit":
                return True
            return False
    return False



# IMPROVED ADD AND REMOVE SO IT TRACKS IN LOG MOVEMENTS
def add_new(product_id, name, category, price, qty, reorder_level):
    ws1.append([product_id.upper(), name.title(), category.title(), safe_float(price), safe_int(qty), safe_int(reorder_level)])
    wb1.save(Inventory_Database)
    log_movement(product_id.upper(), "IN", qty, "Initial stock")
    return True

def remove_product(product_id):
    for row in ws1.iter_rows(min_row=2):
        if str(row[0].value).strip().upper() == product_id.upper():
            ws1.delete_rows(row[0].row, 1)
            wb1.save(Inventory_Database)
            log_movement(product_id, "REMOVE", 0, "Product removed from inventory")
            return True
    return False


# CHANGE get_price(code) TO GET_PRICE(PRODUCT_ID) FOR CONSISTENCY
def get_price(product_id):
    for row in ws1.iter_rows(min_row=2, values_only=True):
        if str(row[0]).strip().upper() == product_id.upper():
            return safe_float(row[3])
    return None

# CHANGE get_product(code) TO GET_PRODUCT(PRODUCT_ID) FOR CONSISTENCY
def get_product(product_id):
    for row in ws1.iter_rows(min_row=2, max_col=1, values_only=True):
        if str(row[0]).strip().upper() == product_id.upper():
            return True
    return False


# MOVED "NOT ENOUGH STOCK" MESSAGE TO buy() AND ADDED LOG MOVEMENT
def update_stock(product_id, qty_sold):
    for row in ws1.iter_rows(min_row=2):
        if str(row[0].value).strip().upper() == product_id.upper():
            current_stock = safe_int(row[4].value)
            new_stock = current_stock - safe_int(qty_sold)
            if new_stock < 0:
                return False, current_stock  
            row[4].value = new_stock
            wb1.save(Inventory_Database)
            log_movement(product_id, "OUT", qty_sold, "Stock adjustment")
            return True, new_stock
    return False, 0


# ADDED AN UNSUCCESSFUL MESSAGE IF STOCK IS INSUFFICIENT
def buy(product_id, qty, sales_id):
    success, new_stock = update_stock(product_id, qty)
    if not success:
        return False, f"Not enough stock. Current stock: {new_stock}"
    
    price = get_price(product_id)
    subtotal = safe_float(price) * safe_int(qty)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws2.append([sales_id, now, product_id, qty, price, subtotal])
    wb2.save(Sales_Database)
    log_movement(product_id, "SALE", qty, "Customer purchase")
    return True, subtotal


# IMPROVED THE PRINT RECEIPT FOR BETTER READABILITY
def print_receipt(sales_id):
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[0] == safe_int(sales_id):
            return {"product_id": row[2], "quantity": row[3], "unit_price": row[4], "total": row[5]}
    return None

# NEW FUNCTION
def log_movement(product_id, movement_type, quantity, remarks):
    movement_id = ws4.max_row
    date_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws4.append([movement_id, product_id, movement_type, quantity, date_now, remarks])
    wb4.save(Movement_Database)

# IMPROVED SAVE FUNCTION, CALL THIS WHENEVER THE SYSTEM IS CLOSED
def save():
    wb1.save("Database.xlsx")
    wb2.save("sale.xlsx")
    wb3.save("user.xlsx")
    wb4.save("inventory_movements.xlsx")

# ----------------- LIST / REPORT FUNCTIONS -----------------
def get_all_products():
    data = []
    for row in ws1.iter_rows(min_row=2, values_only=True):
        data.append({
            "product_id": row[0] or "-",
            "name": row[1] or "-",
            "category": row[2] or "-",
            "price": safe_float(row[3]),
            "stock": safe_int(row[4]),
            "reorder": safe_int(row[5])
        })
    return data

def get_all_sales():
    data = []
    for row in ws2.iter_rows(min_row=2, values_only=True):
        data.append({
            "sale_id": row[0] or "-",
            "date": row[1] or "-",
            "product_id": row[2] or "-",
            "quantity": safe_int(row[3]),
            "unit_price": safe_float(row[4]),
            "total": safe_float(row[5])
        })
    return data

def get_inventory_movements():
    data = []
    for row in ws4.iter_rows(min_row=2, values_only=True):
        data.append({
            "movement_id": row[0] or "-",
            "product_id": row[1] or "-",
            "type": row[2] or "-",
            "quantity": safe_int(row[3]),
            "date": row[4] or "-",
            "remarks": row[5] or "-"
        })
    return data

def sales_summary():
    total_sales = total_items = total_revenue = 0
    for row in ws2.iter_rows(min_row=2, values_only=True):
        total_sales += 1 if row[0] else 0
        total_items += safe_int(row[3])
        total_revenue += safe_float(row[5])
    return {"total_sales": total_sales, "total_items": total_items, "total_revenue": total_revenue}

def best_selling_products(top_n=10):
    product_sales = {}
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[0]:
            pid = row[2]
            qty = safe_int(row[3])
            product_sales[pid] = product_sales.get(pid, 0) + qty
    sorted_sales = sorted(product_sales.items(), key=lambda x: x[1], reverse=True)
    data = []
    for pid, total in sorted_sales[:top_n]:
        data.append({
            "product_id": pid,
            "name": product_name(pid),
            "total_sold": total
        })
    return data

def low_stock_alerts():
    data = []
    for row in ws1.iter_rows(min_row=2, values_only=True):
        pid = row[0]
        name = row[1]
        stock = safe_int(row[4])
        reorder = safe_int(row[5])
        if stock <= reorder:
            data.append({"product_id": pid, "name": name, "stock": stock, "reorder": reorder})
    return data

def remove_ppl(removed_name):
    for row in ws3.iter_rows(min_row=1):
        if str(row[0].value).strip().upper() == removed_name:
            ws1.delete_rows(row[0].row, 1)
            wb1.save(Inventory_Database)
            return True
    return False
def add_ppl(username,role,password):
    ws3.append([username,role,password])
    wb3.save(Inventory_Database)
    return True