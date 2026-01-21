from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import Backend
import random
import getpass
from openpyxl import load_workbook, Workbook

POS = Tk()
POS.geometry("1100x600+150+10")
loginframe = Frame(POS, background='#275c70')
salesframe = Frame(POS, background="#4D6E55")

Sales_ID = random.randint(1000, 9999)
Temp_Total = 0.0
total_var = StringVar()
total_var.set(f"Total:{Temp_Total:.2f} ")
change = 0.0
is_admin = False
is_audit = False

# ----------------------LOGIN FRAME-------------------------------------------------------------------
for i in range(11):
    loginframe.rowconfigure(i, weight=1)
for i in range(10):
    loginframe.columnconfigure(i, weight=2)

label = Label(loginframe,
              text="POSSYS SOLUTIONS",
              font=("times new roman", 24, 'bold'),
              bg="#00403d",
              fg='#B8E3E9',
              relief=RIDGE,
              bd='15',
              width=20)
label.place(relx=0.52, rely=0.1, width=600, anchor='center')

label = Label(loginframe,
              text="Final Capstone Project: Sales and Inventory Management System",
              font=("times new roman", 14, 'bold'),
              bg="#00403d",
              fg='#B8E3E9',
              relief=RIDGE,
              bd='15',
              width=20)
label.place(relx=0.52, rely=0.2, width=600, anchor='center')

user_entry = Entry(loginframe, font=('Arial', 20, 'bold'), width=30, )
Label(loginframe, text="USERNAME",
      font=("times new roman", 17),
      bg="#00403d",
      fg='#B8E3E9',
      relief=RIDGE,
      bd='10').grid(row=6, column=4)
user_entry.grid(row=6, column=5, sticky="ew")

Label(loginframe, text="Password",
      font=("times new roman", 19),
      bg="#00403d",
      fg='#B8E3E9',
      relief=RIDGE,
      bd='10',
      width=9).grid(row=7, column=4)
user_password = Entry(loginframe, font=('Arial', 20, 'bold'), width=30, show='*')
user_password.grid(row=7, column=5, sticky="ew")


def login_user():
    global is_admin
    global is_audit
    if Backend.login(user_entry.get(), user_password.get()):
        if Backend.checkadmin(user_entry.get()):
            is_admin = True
        elif Backend.checkaudit(user_entry.get()):
            is_audit = True
        user_entry.delete(0, END)
        user_password.delete(0, END)
        loginframe.pack_forget()
        salesframe.pack(fill="both", expand=True)
    else:
        warn = Label(loginframe, text="!INVALID Username or Password!",
                     font=("times new roman", 17, 'bold'),
                     bg="#00403d",
                     fg='#B8E3E9',
                     relief=RAISED,
                     bd='5',
                     width=26)
        warn.grid(row=11, column=5)


Button(loginframe, text='Login', font=("times new roman", 17, 'bold'),
       bg="#00403d",
       fg='#B8E3E9',
       activebackground='#4bafd6',
       activeforeground='white',
       relief=GROOVE,
       bd='10', command=login_user).grid(row=9, column=5, sticky="news")
loginframe.pack(fill="both", expand=True)
devs = Label(loginframe,
             text='Developed by: \n-JUSTINE BONGALOS,\n-JOAQUIN DORONGON,\n-MARK OCAMPO,\n -MJ DAMIAN',
             font=('Arial', 10, 'bold'),
             bg="#00403d",
             fg='#B8E3E9',
             relief=SOLID,
             bd=5)
devs.grid(row=11, column=9)

# ------------------------MAIN POS FRAME-----------------------------------------------------
for i in range(10):
    salesframe.rowconfigure(i, weight=1)
for i in range(7):
    salesframe.columnconfigure(i, weight=1)
Label(salesframe, text="POSSYS SOLUTIONS",
      font=('times new roman', 24, 'bold'),
      bg="#00403d",
      fg='#B8E3E9',
      relief=RIDGE,
      bd=10).grid(row=0, column=4, columnspan=1,
                  sticky='nswe')

product_entry = Entry(salesframe)
Label(salesframe, text="ENTRY >>>", font=("Times new Roman", 15, "bold")).grid(row=1, column=1, sticky="we")
product_entry.grid(row=1, column=3, columnspan=3, sticky="we")
item_list = ttk.Treeview(salesframe, columns=("Product", "Quantity", "Price", "Subtotal"), show="headings", height=10)
item_list.grid(row=2, column=3, columnspan=3, rowspan=7, sticky='nsew')
item_list.heading("Product", text="Product")
item_list.heading("Quantity", text="Quantity")
item_list.heading("Price", text="Price")
item_list.heading("Subtotal", text="Subtotal")


def addtocart(event=None):
    global Temp_Total
    product = product_entry.get()
    if "*" in product_entry.get():
        try:
            product_code, qty = product.split("*")
            product_code = product_code.strip()
            qty = int(qty.strip())
        except:
            messagebox.showerror("Invalid entry format!", "Use: Code * Quantity")
            product_entry.delete(0, END)
            return
    else:
        product_code = product
        qty = 1

    if Backend.get_product(product_code):
        Backend.buy(product_code, qty, Sales_ID)
        item_list.insert("", END, values=(Backend.product_name(product_code), qty, Backend.get_price(product_code),
                                          float(qty * Backend.get_price(product_code))))
        product_entry.delete(0, END)
        Temp_Total += float(qty * Backend.get_price(product_code))
        total_var.set(f"Total: {Temp_Total:.2f}")
    else:
        messagebox.showerror("WARNING", "INVALID PRODUCT ID")
        product_entry.delete(0, END)


def finish_transaction():
    popup = Toplevel(POS)
    popup.title("Finish Transaction")
    popup.geometry("500x150+300+150")
    popup.config(background='#275c70')
    popup.grab_set()
    Label(popup, text=f"TOTAL: {Temp_Total}", font=("times new roman", 15, 'bold'),
          bg="#00403d",
          fg='#B8E3E9',
          relief=RIDGE,
          bd='5',
          width=20).pack()
    Label(popup, text="Cash Recieved:", font=("times new roman", 15, 'bold'),
          bg="#00403d",
          fg='#B8E3E9',
          relief=RIDGE,
          bd='5',
          width=20).pack()
    cash = Entry(popup, font=("arial", 15, 'bold'), relief=GROOVE, bd=5)
    cash.pack()

    def changes(event=None):
        global Temp_Total
        global change
        global Sales_ID
        try:
            change += (float(cash.get()) - Temp_Total)
            messagebox.showinfo("CHANGE", f" Your Change is: {change}")
            Sales_ID = random.randint(1000, 9999)
            for item in item_list.get_children():
                item_list.delete(item)
            Temp_Total = 0
            total_var.set(f"Total: {Temp_Total:.2f} ")
            Backend.save()
            popup.destroy()


        except ValueError:
            messagebox.showerror("Invalid entry format!", "YOU CANNOT SUBTRACT THAT TO A NUMBER YEA?")
            cash.delete(0, END)

    Button(popup, text="Confirm", font=('arial', 15, 'bold'),
               bg="#1a7000",
               fg='black',
               activebackground='#33d900',
               activeforeground='white',
           relief=GROOVE,
           bd=10, width=15, command=changes).pack()


def printreceipt():
    popup = Toplevel(POS)
    popup.title("PRINT RECEIPT")
    popup.geometry("700x500+300+150")
    popup.config(background='#275c70')
    popup.grab_set()
    Label(popup, text="Enter Trasaction ID",font=("times new roman", 15, 'bold'),
          bg="#00403d",
          fg='#B8E3E9',
          relief=RIDGE,
          bd='5',
          width=20).pack()
    transac_id = Entry(popup,  font=('times new roman', 15, 'bold'),
               bg="#1a7000",
               fg='black',
                relief=GROOVE,
                 bd=10, width=15)
    transac_id.pack()
    item_list = ttk.Treeview(popup, columns=("Product", "Quantity", "Price", "Subtotal"), show="headings", height=10)
    item_list.pack()

    def printrec():
        wb2 = load_workbook("sale.xlsx")
        ws2 = wb2.active
        for row in ws2.iter_rows(min_row=1, values_only=True):
            if row[0] == int(transac_id.get().strip()):
                item_list.insert("", END, values=(row[2], row[3], row[4], row[5]))

    def close():
        popup.destroy()

    Button(popup, text="Submit",
           font=("times new roman", 20, 'bold'),
           bg="#00403d",
           fg='#B8E3E9',
           relief=RIDGE,
           bd='5',
           width=20,
           command=printrec).pack()
    Button(popup, text="Close",
           font=('times new roman', 20, 'bold'),
           bg="red",
           fg='white',
           activebackground='maroon',
           activeforeground='white',
           relief=GROOVE,
           bd=10, width=15,
           command=close).pack()


def logout():
    pop = Toplevel(POS, background='#940000')
    pop.title("WARNING")
    pop.geometry("300x200+500+250")
    Label(pop,text="LOG OUT?",font=("times new roman",20,"bold"),bg='Yellow', fg='Black',relief=RIDGE, bd=10).place(x=70, y=10)
    def yeah():
        pop.destroy()
        salesframe.pack_forget()
        loginframe.pack(fill = "both",expand=True)
    def nah():
        pop.destroy()
    Button(pop,text="Yes",
      font=('times new roman',15,'bold'),
      bg="#1a7000",
      fg='black',
      activebackground='#33d900',
      activeforeground='white',
      relief=RIDGE,
      bd=10,command=yeah).place(x=122, y=70)
    Button(pop,text="Nah",
      font=('times new roman',15,'bold'),
      bg="#991003",
      fg='black',
      activebackground='#e31502',
      activeforeground='white',
      relief=RIDGE,
      bd=10,command=nah).place(x=120, y=130)

def change_stock():
    if is_audit or is_admin:
        Inventory_Database = "Database.xlsx"
        wb1 = load_workbook(Inventory_Database)
        ws1 = wb1.active
        popup = Toplevel(POS)
        popup.title("Change stock")
        popup.geometry("700x500+300+150")
        popup.config(background='#275c70')
        popup.grab_set()
        Label(popup, text="Low Stocks", font=("times new roman", 25, 'bold'),
          bg="#00403d",
          fg='#B8E3E9',
          relief=RIDGE,
          bd='5',
          width=20).pack()
        stocks_table = ttk.Treeview(popup, columns=("id", "name", "stock", "order"), show="headings")
        stocks_table.heading("id", text="Rank")
        stocks_table.heading("name", text="Product Name")
        stocks_table.heading("stock", text="Stock/s")
        stocks_table.heading("order", text="ORDER")
        stocks_table.column("id", width=60, anchor=CENTER)
        stocks_table.column("name", width=100)
        stocks_table.column("stock", width=200, anchor=W)
        stocks_table.column("order", width=100, anchor=CENTER)
        stocks = Backend.low_stock_alerts()
        for products in stocks:
            if products["product_id"] != None:
                stocks_table.insert(
                    "",
                    END,
                    values=(
                        products["product_id"],
                        products["name"],
                        products["stock"],
                        products["reorder"]
                    )
                )
        stocks_table.pack()
        Label(popup, text="Enter Product ID", font=("times new roman", 15, 'bold'),
          bg="#00403d",
          fg='#B8E3E9',
          relief=SOLID,
          bd='5',
          width=20).pack()
        product_entry = Entry(popup, font=("arial", 15, 'bold'), relief=GROOVE, bd=5)
        product_entry.pack()
        Label(popup, text="Enter New Stock", font=("times new roman", 15, 'bold'),
          bg="#00403d",
          fg='#B8E3E9',
          relief=SOLID,
          bd='5',
          width=20).pack()
        stock_entry = Entry(popup, font=("arial", 15, 'bold'), relief=GROOVE, bd=5)
        stock_entry.pack()

        def confirmstock():
            if Backend.get_product(product_entry.get().strip().upper()):
                Backend.change_stock(product_entry.get().strip(), stock_entry.get().strip())
                messagebox.showinfo("STOCKS", "Stocks has been updated")
                popup.destroy()
            else:
                messagebox.showwarning("WARNING", "ID NOT FOUND")

        def close():
            popup.destroy()

        Button(popup, text="Submit",font=('arial', 10, 'bold'),
           bg="#1a7000",
           fg='white',
           activebackground='#33d900',
           activeforeground='white',
           relief=GROOVE,
           bd=5, width=15, command=confirmstock).pack()
        Button(popup, text="Cancel", font=('arial', 10, 'bold'),
           bg="red",
           fg='white',
           activebackground='maroon',
           activeforeground='white',
           relief=GROOVE,
           bd=5, width=15,command=close).pack()
    else:
        messagebox.showwarning("WARNING", "You Need to be an AUDIT/ADMIN to make changes")


def add_product():
    popup = Toplevel(POS)
    popup.title("Add Product")
    popup.geometry("700x500+300+150")
    popup.config(background='#275c70')
    popup.grab_set()
    Label(popup, text="Enter Product ID", font=("times new roman", 15, 'bold'),
          bg="#00403d",
          fg='#B8E3E9',
          relief=RIDGE,
          bd='5',
          width=20).pack()
    product = Entry(popup, font=("arial", 15, 'bold'))
    product.pack()
    Label(popup, text="Enter Product Name", font=("times new roman", 15, 'bold'),
          bg="#00403d",
          fg='#B8E3E9',
          relief=RIDGE,
          bd='5',
          width=20).pack()
    name = Entry(popup, font=("arial", 15, 'bold'))
    name.pack()
    Label(popup, text="Enter Product Category", font=("times new roman", 15, 'bold'),
          bg="#00403d",
          fg='#B8E3E9',
          relief=RIDGE,
          bd='5',
          width=20).pack()
    category = Entry(popup, font=("arial", 15, 'bold'))
    category.pack()
    Label(popup, text="Enter Product Price", font=("times new roman", 15, 'bold'),
          bg="#00403d",
          fg='#B8E3E9',
          relief=RIDGE,
          bd='5',
          width=20).pack()
    price = Entry(popup, font=("arial", 15, 'bold'))
    price.pack()
    Label(popup, text="Enter Product Stock/s", font=("times new roman", 15, 'bold'),
          bg="#00403d",
          fg='#B8E3E9',
          relief=RIDGE,
          bd='5',
          width=20).pack()
    stocks = Entry(popup, font=("arial", 15, 'bold'))
    stocks.pack()
    Label(popup, text="Enter Product Reorder lvl", font=("times new roman", 15, 'bold'),
          bg="#00403d",
          fg='#B8E3E9',
          relief=RIDGE,
          bd='5',
          width=20).pack()
    reorder = Entry(popup, font=("arial", 15, 'bold'))
    reorder.pack()

    def submit():
        try:
            Backend.add_new(product.get(), name.get(), category.get(), price.get(), stocks.get(), reorder.get())
            messagebox.showinfo("SUCESS",
                                f"Successfully added {name.get()} worth {price.get()} with {stocks.get()} stocks")
            product.delete(0, END)
            name.delete(0, END)
            category.delete(0, END)
            price.delete(0, END)
            stocks.delete(0, END)
            reorder.delete(0, END)

        except ValueError:
            messagebox.showwarning("Warning", "pls enter integers/decimals on price,stocks and reorder lvl")
            price.delete(0, END)
            stocks.delete(0, END)
            reorder.delete(0, END)

    def exit():
        popup.destroy()

    Button(popup, text="Submit", font=('arial', 15, 'bold'),
           bg="#1a7000",
           fg='black',
           activebackground='#33d900',
           activeforeground='white',
           relief=GROOVE,
           bd=10, width=15,
           command=submit).pack()
    Button(popup, text="Exit", font=('arial', 15, 'bold'),
           bg="red",
           fg='white',
           activebackground='maroon',
           activeforeground='white',
           relief=GROOVE,
           bd=10, width=15,
           command=exit).pack()


def remove():
    if is_admin or is_audit:
        popup = Toplevel(POS)
        popup.title("Remove Product")
        popup.geometry("700x220+300+150")
        popup.config(background='#275c70')
        popup.grab_set()
        lebel = Label(text='')
        label.pack
        Label(popup, text="Enter Product ID to be Removed",
              font=("times new roman", 25, 'bold'),
          bg="#00403d",
          fg='#B8E3E9',
          relief=RIDGE,
          bd='5',
          width=40).pack()
        product = Entry(popup,font=('arial', 15, 'bold'),
                        relief=SOLID,bd=5)
        product.pack()

        def submit():
            if Backend.get_product(product.get()):
                pop = Toplevel(POS)
                pop.title("Confirm Removal")
                pop.geometry("250x150")
                pop.grab_set()
                Label(pop, text=f"Are you sure you wanna remove {Backend.product_name(product.get())}?").pack()

                def ye():
                    pop.destroy()
                    messagebox.showinfo("SUCCESS", f"Sucessfully removed {Backend.product_name(product.get())} ")
                    Backend.remove(product.get())
                    product.delete(0, END)

                def nah():
                    pop.destroy()

                Button(pop, text="YE", command=ye).pack()
                Button(pop, text="NAH", command=nah).pack()
            else:
                messagebox.showwarning("W", "Cant Find Code")

        def cancel():
            popup.destroy()

        Button(popup, text="Submit", font=('arial', 15, 'bold'),
               bg="#1a7000",
               fg='black',
               activebackground='#33d900',
               activeforeground='white',
           relief=GROOVE,
           bd=10, width=15, command=submit).pack()
        Button(popup, text="Cancel",  font=('arial', 15, 'bold'),
           bg="red",
           fg='white',
           activebackground='maroon',
           activeforeground='white',
           relief=GROOVE,
           bd=10, width=15,command=cancel).pack()
    else:
        messagebox.showwarning("WARNING", "You Need to be an AUDIT/ADMIN to make changes")


def sales_summary():
    popup = Toplevel(POS)
    popup.title("ANALYTICS")
    popup.geometry("800x550+330+70")
    popup.config(background='#3e5944')
    popup.grab_set()
    Label(popup, text="PRODUCT RANKINGS BASED ON NO# of SALES", font=("times new roman", 16, "bold"),bg="#00403d",
                    fg='#B8E3E9',
                    relief=GROOVE,
                    bd = '5').pack()
    rank_table = ttk.Treeview(popup, columns=("rank", "id", "product", "sold"), show="headings")
    rank_table.heading("rank", text="Rank")
    rank_table.heading("product", text="Product Name")
    rank_table.heading("sold", text="Units Sold")
    rank_table.heading("id", text="Product ID")
    rank_table.column("rank", width=60, anchor=CENTER)
    rank_table.column("id", width=100)
    rank_table.column("product", width=200, anchor=W)
    rank_table.column("sold", width=100, anchor=CENTER)
    rankings = Backend.best_selling_products()
    rank_table.pack()
    for rank, item in enumerate(rankings, start=1):
        rank_table.insert(
            "",
            "end",
            values=(
                rank,
                item["product_id"],
                item["name"],
                item["total_sold"])
        )
    Label(popup, text='AVAILABLE STOCKS',font=("times new roman", 16, "bold"),bg="#00403d",
                    fg='#B8E3E9',
                    relief=GROOVE,
                    bd = '5').pack()
    product = Backend.get_all_products()
    product_list = ttk.Treeview(
        popup,
        columns=("id", "name", "category", "price", "stock", "reorder"),
        show="headings",
        height=10
    )
    product_list.heading("id", text="ID")
    product_list.heading("name", text="Name")
    product_list.heading("category", text="Category")
    product_list.heading("price", text="Price")
    product_list.heading("stock", text="Stock")
    product_list.heading("reorder", text="Reorder Level")
    product_list.column("id", width=60, anchor="center")
    product_list.column("name", width=150)
    product_list.column("category", width=120)
    product_list.column("price", width=80, anchor="e")
    product_list.column("stock", width=80, anchor="center")
    product_list.column("reorder", width=100, anchor="center")

    for p in product:
        product_list.insert(
            "",
            "end",
            values=(
                p["product_id"],
                p["name"],
                p["category"],
                p["price"],
                p["stock"],
                p["reorder"]
            )
        )

    product_list.pack()


def remove_user():
    if is_admin:
        popup = Toplevel(POS)
        popup.title("Remove User")
        popup.geometry("450x450+370+150")
        popup.config(background='#3e5944')
        popup.grab_set()
        Label(popup, text="Enter Username to be Removed",font=("arial",18),
                    bg="#00403d",
                    fg='#B8E3E9',
                    relief=GROOVE,
                    bd = '5').pack()
        user = Entry(popup, font=('arial', 20, 'bold'), relief=SOLID, bd=5)
        user.pack()

        def submit():
            if Backend.check_ppl(user.get()):
                pop = Toplevel(POS)
                pop.title("Confirm Removal")
                pop.geometry("250x150")
                pop.grab_set()
                Label(pop, text=f"Are you sure you wanna remove {user.get()}?").pack()

                def ye():
                    pop.destroy()
                    messagebox.showinfo("SUCCESS", f"Sucessfully removed {user.get()} ")
                    Backend.remove_ppl(user.get())
                    user.delete(0, END)

                def nah():
                    pop.destroy()
            else:
                messagebox.showwarning("Warning",f"{user.get()} ,is an invalid username")
            Button(pop, text="YE", command=ye).pack()
            Button(pop, text="NAH", command=nah).pack()

        def cancel():
            popup.destroy()

        Button(popup, text="Submit", font = ('arial', 30,'bold'),
                            bg="#00403d",
                            fg='#B8E3E9',
                            activebackground='#4bafd6',
                            activeforeground='white',
                            relief=GROOVE,
                            bd='5', command=submit).pack()
        Button(popup, text="Cancel", font = ('arial', 30,'bold'),
                            bg="#00403d",
                            fg='#B8E3E9',
                            activebackground='#4bafd6',
                            activeforeground='white',
                            relief=GROOVE,
                            bd='5', command=cancel).pack()
    else:
        messagebox.showwarning("WARNING", "You Need to be an ADMIN to make changes")


def add_user():
    if is_admin:
        roles = ["Admin","Audit","Cashier"]
        selected_role = StringVar()
        selected_role.set(roles[0])
        popup = Toplevel(POS)
        popup.title("Add People")
        popup.geometry("700x500+330+120")
        popup.config(background='#3e5944')
        popup.grab_set()
        Label(popup,text="Enter Username",font=("arial",18),
                    bg="#00403d",
                    fg='#B8E3E9',
                    relief=GROOVE,
                    bd = '5').pack()
        user = Entry(popup, font=('arial', 20, 'bold'), relief=SOLID, bd=5)
        user.pack()
        Label(popup,text="Enter Password",font=("arial",18),
                   bg="#00403d",
                   fg='#B8E3E9',
                   relief=GROOVE,
                   bd = '5').pack()
        password = Entry(popup, font=('arial', 20, 'bold'), relief=SOLID, bd=5)
        password.pack()
        role_menu = OptionMenu(popup, selected_role, *roles)
        role_menu.pack(pady=20)

        def confirm():
            Backend.add_ppl(user.get(), selected_role.get(), password.get())
            messagebox.showinfo("SUCCESS", f"{user.get()} is sucessfully added as {selected_role.get()}")
            user.delete(0,END)
            password.delete(0,END)

        def cancel():
            popup.destroy()

        Button(popup, text="Add",bg="#00403d", font = ('arial', 30,'bold'),
                            fg='#B8E3E9',
                            activebackground='#4bafd6',
                            activeforeground='white',
                            relief=GROOVE,
                            bd='5', command=confirm).pack()
        Button(popup, text="Cancel", font = ('arial', 30,'bold'),
                            bg="#00403d",
                            fg='#B8E3E9',
                            activebackground='#4bafd6',
                            activeforeground='white',
                            relief=GROOVE,
                            bd='5', command=cancel)
    else:
        messagebox.showwarning("WARNING", "You Need to be an ADMIN to make changes")


product_entry.bind("<Return>", addtocart)
Label(salesframe, textvariable=total_var, font=("arial", 18),
      bg="#00403d",
      fg='#B8E3E9',
      relief=GROOVE,
      bd='5').grid(column=4, row=9, sticky="nsew", pady=10)
logout_button = Button(salesframe, text="Log Out", font=("arial", 13), command=logout,
                       bg="#00403d",
                       fg='#B8E3E9',
                       activebackground='#4bafd6',
                       activeforeground='white',
                       relief=GROOVE,
                       bd='5')
logout_button.grid(row=2, column=0, sticky="nsew", columnspan=2)

add_product_button = Button(salesframe, text="Add Product", font=("arial", 13), command=add_product,
                            bg="#00403d",
                            fg='#B8E3E9',
                            activebackground='#4bafd6',
                            activeforeground='white',
                            relief=GROOVE,
                            bd='5')
add_product_button.grid(row=3, column=0, sticky="nsew")

remove_product_button = Button(salesframe, text="Remove Product", font=("arial", 13), command=remove,
                               bg="#00403d",
                               fg='#B8E3E9',
                               activebackground='#4bafd6',
                               activeforeground='white',
                               relief=GROOVE,
                               bd='5')
remove_product_button.grid(row=4, column=0, sticky="nsew")

change_stock_button = Button(salesframe, text="Change Stock", font=("arial", 13), command=change_stock,
                             bg="#00403d",
                             fg='#B8E3E9',
                             activebackground='#4bafd6',
                             activeforeground='white',
                             relief=GROOVE,
                             bd='5')
change_stock_button.grid(row=5, column=0, sticky='nsew')

print_reciept_button = Button(salesframe, text="Print reciept", font=("arial", 13), command=printreceipt,
                              bg="#00403d",
                              fg='#B8E3E9',
                              activebackground='#4bafd6',
                              activeforeground='white',
                              relief=GROOVE,
                              bd='5')
print_reciept_button.grid(row=6, column=0, sticky='nsew')

sale_summary_button = Button(salesframe, text="Sales Summary", command=sales_summary, font=("arial", 13),
                             bg="#00403d",
                             fg='#B8E3E9',
                             activebackground='#4bafd6',
                             activeforeground='white',
                             relief=GROOVE,
                             bd='5')
sale_summary_button.grid(row=3, column=1, sticky='nsew')

add_user_button = Button(salesframe, text="Add User", command=add_user, font=("arial", 13),
                         bg="#00403d",
                         fg='#B8E3E9',
                         activebackground='#4bafd6',
                         activeforeground='white',
                         relief=GROOVE,
                         bd='5')
add_user_button.grid(row=4, column=1, sticky='nsew')

remove_user_button = Button(salesframe, text="Remove User", command=remove_user, font=("arial", 13),
                            bg="#00403d",
                            fg='#B8E3E9',
                            activebackground='#4bafd6',
                            activeforeground='white',
                            relief=GROOVE,
                            bd='5')
remove_user_button.grid(row=5, column=1, sticky='nsew')

finish_transac_button = Button(salesframe, text="Finish Transaction", font=("arial", 13), command=finish_transaction,
                               bg="#00403d",
                               fg='#B8E3E9',
                               activebackground='#4bafd6',
                               activeforeground='white',
                               relief=GROOVE,
                               bd='5')
finish_transac_button.grid(row=6, column=1, sticky="nsew")

POS.mainloop()