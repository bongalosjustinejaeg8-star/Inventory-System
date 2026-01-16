from openpyxl import load_workbook, Workbook
import random
from datetime import datetime
Inventory_Database = "Database.xlsx"
Sales_Database = "sale.xlsx"

wb1 = load_workbook(Inventory_Database)
ws1 = wb1.active
wb2 = load_workbook(Sales_Database)
ws2 = wb2.active



def chkproduct(code):
    for cell in ws1.iter_rows(min_row=1, max_col=1, values_only=True):
        if cell[0] == code:
            return True
    return False

def price(code):
    for row in ws1.iter_rows(min_row=1, max_col=5, values_only=True):
        if row[0] == code:
            return row[3]
    return None
def printreciept(sales_id):
    print("Reciept")
    print(f"Sales ID: {sales_id}")
    for row in ws2.iter_rows(min_row=1,values_only=True):
        if row[0] == sales_id:
            print(row[2], row[3], row[4], row[5])
    print("-"*10)
def update_stock(code, qty_sold):
    for row in ws1.iter_rows(min_row=2):
        if str(row[0].value).strip().upper() == code:
            current_stock = row[4].value
            new_stock = current_stock - qty_sold
            if new_stock < 0:
                print(f"Not enough stock for {code}! Only {current_stock} left.")
                return False
            row[4].value = new_stock
            wb1.save(Inventory_Database)
            return True
    return False 
def buy():
    now = datetime.now()
    sales_id = random.randint(10000, 99999)

    while True:
        sale_input = input("Enter product code*qty or X to finish: ").upper()
        if sale_input == "X":
            break

        if "*" in sale_input:
            code, qty = sale_input.split("*")
            qty = int(qty)
        else:
            code = sale_input
            qty = 1

        if not chkproduct(code):
            input("Item not found. Press Enter to continue...")
            continue
        
        update_stock(code,qty)
        p = price(code)
        subtotal = float(p) * qty
        ws2.append([sales_id, now.strftime("%Y-%m-%d %H:%M:%S"), code, qty, p, float(subtotal)])

    wb2.save("sale.xlsx")
    wb1.save("Database.xlsx")
    printreciept(sales_id)
    print("Sale complete!")
