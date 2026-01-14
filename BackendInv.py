from openpyxl import load_workbook
file=load_workbook("DatabaseCap.xlsx")
file2 = file.active


def add_new():
    file2.append([input("enter product name: "),int(input("enter stocks: "))])
    file.save("DatabaseCap.xlsx")
    


def change_stock():
    product_name = input("enter product name: ")
    found=False
    for rows in file2.iter_rows(min_row=2):
        if rows[0].value == product_name:
            found=True
            rows[1].value = int(input("enter new quantity: "))
            file.save("DatabaseCap.xlsx")
            print ("stock updated")
            
            break
    if not found:
        print(f"{product_name} not found")    

def remv_item():
    item = input("Enter Item ID to remove: ").title()
    found = False

    for row in file2.iter_rows(min_row=2):
        if row[0].value== item:
            found = True
            decision = input(
                f"Are you sure you want to remove {row[0].value}? (y/n): "
            ).lower()

            if decision == "y":
                row[0].value = None
                row[1].value = None
                file.save("DatabaseCap.xlsx")
                print("Item removed successfully.")
            else:
                print("Removal cancelled.")
            break

    if not found:
        dec = input(f"{item} not found, press r to retry").lower()
        if dec == "r":
            remv_item()

for row,row1 in file2.iter_rows(min_row=2,values_only=1):
    if row==None:
        continue
    print(row,row1)


