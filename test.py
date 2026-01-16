from tkinter import *
import Sales
from openpyxl import load_workbook, Workbook
import random
from datetime import datetime
Inventory_Database = "Database.xlsx"
Sales_Database = "sale.xlsx"

wb1 = load_workbook(Inventory_Database)
ws1 = wb1.active
wb2 = load_workbook(Sales_Database)
ws2 = wb2.active

def buy(sale_input):
    now = datetime.now()
    sales_id = random.randint(10000, 99999)

    while True:
        if sale_input == "X":
            break

        if "*" in sale_input:
            code, qty = sale_input.split("*")
            qty = int(qty)
        else:
            code = sale_input
            qty = 1

        if not Sales.chkproduct(code):
            input("Item not found. Press Enter to continue...")
            continue
        
        Sales.update_stock(code,qty)
        p = Sales.price(code)
        subtotal = float(p) * qty
        ws2.append([sales_id, now.strftime("%Y-%m-%d %H:%M:%S"), code, qty, p, float(subtotal)])

    wb2.save("sale.xlsx")
    wb1.save("Database.xlsx")
    Sales.printreciept(sales_id)
    print("Sale complete!")



Main_window = Tk()
Main_window.geometry("400x200")
product_entry = Entry(Main_window)
product_entry.pack()
submit_button = Button(text="Submit",command=submit).pack()



















Main_window.mainloop()